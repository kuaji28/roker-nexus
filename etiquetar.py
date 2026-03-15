"""
ETIQUETAR.PY — Identificador de exports Flexxus
================================================
El problema: Flexxus siempre genera el mismo nombre de archivo
sin importar qué depósito exportaste. El archivo tampoco dice
de qué depósito es adentro.

La solución: este script
  1. Detecta automáticamente los xlsx recientes en Descargas
  2. Te pregunta qué depósito es
  3. Renombra el archivo con el código correcto
  4. INYECTA el nombre del depósito DENTRO del archivo (celda visible)
     → así el archivo se identifica solo, aunque lo muevas o renombres

USO: doble clic en etiquetar.py  (o: python etiquetar.py)
"""

import os
import sys
import glob
import shutil
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("Instalando dependencia openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

# ─────────────────────────────────────────────
#  CONFIGURACIÓN DE DEPÓSITOS
# ─────────────────────────────────────────────
DEPOSITOS = {
    "1": {"codigo": "SJ",  "nombre": "SAN JOSE",              "full": "DEPOSITO SAN JOSE (hub principal)"},
    "2": {"codigo": "LAR", "nombre": "LARREA",                 "full": "DEPOSITO LARREA (local al público)"},
    "3": {"codigo": "SAR", "nombre": "SARMIENTO",              "full": "DEPOSITO SARMIENTO NUEVO2"},
    "4": {"codigo": "FML", "nombre": "FULL ML",                "full": "DEP. FULL ML (MercadoLibre Fulfillment)"},
    "5": {"codigo": "DML", "nombre": "DEP. MERCADO LIBRE",     "full": "DEPOSITO MERCADO LIBRE"},
    "6": {"codigo": "MER", "nombre": "MERMAS",                 "full": "MERMAS GENERALES"},
    "7": {"codigo": "RMA", "nombre": "DEP. RMA",               "full": "DEP. TRANSITORIO RMA"},
    "8": {"codigo": "MUE", "nombre": "MUESTRAS",               "full": "DEPOSITO MUESTRAS"},
    "9": {"codigo": "UI",  "nombre": "USO INTERNO",            "full": "USO INTERNO"},
}

# Tipos de reporte reconocidos
TIPOS_REPORTE = {
    "1": {"codigo": "stock",      "nombre": "Stock (Planilla de Stock / Listado General)"},
    "2": {"codigo": "historico",  "nombre": "Histórico de Artículos"},
    "3": {"codigo": "ventas",     "nombre": "Venta x Artículo x Mes"},
    "4": {"codigo": "remitos",    "nombre": "Remitos Internos"},
    "5": {"codigo": "rma",        "nombre": "RMA / Devoluciones"},
    "6": {"codigo": "export",     "nombre": "Otro export Flexxus"},
}

# Destino final de archivos etiquetados
SCRIPT_DIR = Path(__file__).parent
EXPORTS_DIR = SCRIPT_DIR / "exports_flexxus"
EXPORTS_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────
#  BUSCAR ARCHIVOS RECIENTES EN DESCARGAS
# ─────────────────────────────────────────────
def buscar_descargas_recientes(minutos=180):
    """Busca xlsx de Flexxus descargados en los últimos N minutos.

    Flexxus genera SIEMPRE estos patrones de nombre:
      • Planilla de Stock_DD-MM-YYYY HH-MM-SS.xlsx       (stock por depósito)
      • Listado Histórico de Artículos_...xlsx            (histórico)
      • Venta x Artículo x Mes_...xlsx                    (ventas mensuales)
      • Listado de Stock por Casilleros_...xlsx            (casilleros)
    Si no hay coincidencia exacta, muestra TODOS los xlsx recientes.
    """
    PATRONES_FLEXXUS = [
        "Planilla de Stock",
        "Listado Hist",          # Histórico de Artículos
        "Venta x Art",           # Venta x Artículo x Mes
        "Listado de Stock",      # por Casilleros
        "Listado Stock",
        "Listado General",
    ]

    posibles_dirs = [
        Path.home() / "Downloads",
        Path.home() / "Descargas",
    ]
    try:
        posibles_dirs.append(Path("C:/Users") / os.getlogin() / "Downloads")
    except Exception:
        pass

    archivos_flexxus = []
    archivos_otros   = []
    ahora  = datetime.now()
    limite = ahora - timedelta(minutes=minutos)

    for carpeta in posibles_dirs:
        if not carpeta.exists():
            continue
        for f in carpeta.glob("*.xlsx"):
            mtime = datetime.fromtimestamp(f.stat().st_mtime)
            if mtime < limite:
                continue
            es_flexxus = any(p.lower() in f.name.lower() for p in PATRONES_FLEXXUS)
            if es_flexxus:
                archivos_flexxus.append((f, mtime))
            else:
                archivos_otros.append((f, mtime))

    # También directorio actual
    for f in Path(".").glob("*.xlsx"):
        mtime = datetime.fromtimestamp(f.stat().st_mtime)
        if mtime > limite:
            archivos_flexxus.append((f, mtime))

    resultado = archivos_flexxus if archivos_flexxus else archivos_otros
    resultado.sort(key=lambda x: x[1], reverse=True)
    return resultado, bool(archivos_flexxus)

# ─────────────────────────────────────────────
#  INYECTAR ETIQUETA DENTRO DEL XLSX
# ─────────────────────────────────────────────
def inyectar_etiqueta(filepath: Path, deposito: dict, tipo: dict, fecha_export: str):
    """Agrega una fila de identificación destacada al inicio del archivo."""
    import stat
    # Asegurar permisos de escritura
    filepath.chmod(stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Insertar 2 filas al principio para no romper la estructura original
    ws.insert_rows(1, 2)

    # ── Fila 1: etiqueta visible ──
    texto = f"▶ DEPÓSITO: {deposito['full']}  |  TIPO: {tipo['nombre']}  |  EXPORTADO: {fecha_export}  |  ID: {deposito['codigo']}"
    ws["A1"] = texto

    # Estilo: fondo amarillo llamativo
    fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    font = Font(bold=True, size=11, color="000000")
    ws["A1"].fill = fill
    ws["A1"].font = font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Fusionar columnas para que se vea completo
    try:
        ws.merge_cells(f"A1:L1")
    except Exception:
        pass

    # ── Fila 2: metadata parseable (para que Nexus v2 lo lea automáticamente) ──
    ws["A2"] = f"NEXUS_META|deposito={deposito['codigo']}|tipo={tipo['codigo']}|fecha={fecha_export}"
    ws["A2"].font = Font(size=7, color="AAAAAA")

    wb.save(filepath)
    print(f"  ✓ Etiqueta inyectada en el archivo")

# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print()
    print("=" * 60)
    print("  ETIQUETAR — Identificador de exports Flexxus")
    print("  EL CELU / Roker Nexus")
    print("=" * 60)

    # 1. Buscar archivos recientes
    print("\n🔍 Buscando exports de Flexxus en Descargas...")
    print(f"   (patrón: 'Planilla de Stock', 'Listado Histórico', etc.)")
    archivos, son_flexxus = buscar_descargas_recientes(minutos=180)

    if not archivos:
        print("\n⚠️  No encontré xlsx recientes en Descargas.")
        print("   Opción A: Arrastrá el archivo sobre esta ventana y presioná Enter")
        print("   Opción B: Escribí la ruta completa del archivo")
        ruta = input("\n📂 Ruta del archivo: ").strip().strip('"')
        if not ruta or not Path(ruta).exists():
            print("❌ Archivo no encontrado. Saliendo.")
            input("\nPresioná Enter para cerrar...")
            return
        archivos = [(Path(ruta), datetime.fromtimestamp(Path(ruta).stat().st_mtime))]
        son_flexxus = False

    if not son_flexxus:
        print("\n   ⚠️  No encontré archivos con nombre típico de Flexxus.")
        print("   Mostrando todos los xlsx recientes:\n")

    # 2. Mostrar lista
    print(f"\n📄 Encontré {len(archivos)} archivo(s) reciente(s):\n")
    for i, (f, mtime) in enumerate(archivos[:10], 1):
        print(f"  {i}. {f.name}")
        print(f"     📁 {f.parent}")
        print(f"     🕐 {mtime.strftime('%d/%m/%Y %H:%M')}\n")

    # 3. Seleccionar archivo
    if len(archivos) == 1:
        idx = 0
        print(f"→ Usando el único archivo encontrado.")
    else:
        eleccion = input("¿Qué número de archivo querés etiquetar? (Enter = el 1er/más reciente): ").strip()
        idx = (int(eleccion) - 1) if eleccion.isdigit() else 0

    archivo_origen = archivos[min(idx, len(archivos)-1)][0]
    fecha_export = datetime.fromtimestamp(archivo_origen.stat().st_mtime).strftime("%Y-%m-%d")

    print(f"\n✅ Archivo seleccionado: {archivo_origen.name}")

    # 4. Seleccionar tipo de reporte
    print("\n📋 ¿QUÉ TIPO DE REPORTE ES?\n")
    for k, v in TIPOS_REPORTE.items():
        print(f"  {k}. {v['nombre']}")
    tipo_input = input("\n¿Número? (Enter = 1, stock): ").strip()
    tipo = TIPOS_REPORTE.get(tipo_input or "1", TIPOS_REPORTE["1"])

    # 5. Seleccionar depósito (solo para reportes de stock/remitos)
    deposito = None
    if tipo["codigo"] in ("stock", "remitos", "export"):
        print("\n🏭 ¿QUÉ DEPÓSITO ES?\n")
        for k, v in DEPOSITOS.items():
            print(f"  {k}. [{v['codigo']}] {v['full']}")
        dep_input = input("\n¿Número?: ").strip()
        deposito = DEPOSITOS.get(dep_input, DEPOSITOS["1"])
    else:
        # Para histórico, ventas, RMA → no aplica depósito específico
        deposito = {"codigo": "ALL", "nombre": "TODOS", "full": "Todos los depósitos"}

    # 6. Construir nombre final
    if deposito["codigo"] != "ALL":
        nuevo_nombre = f"{deposito['codigo']}_{tipo['codigo']}_{fecha_export}.xlsx"
    else:
        nuevo_nombre = f"{tipo['codigo']}_{fecha_export}.xlsx"

    destino = EXPORTS_DIR / nuevo_nombre

    # Evitar sobrescribir: agregar sufijo si ya existe
    counter = 1
    while destino.exists():
        stem = f"{deposito['codigo']}_{tipo['codigo']}_{fecha_export}_v{counter}"
        destino = EXPORTS_DIR / f"{stem}.xlsx"
        counter += 1

    # 7. Copiar + etiquetar
    print(f"\n📝 Procesando...")
    shutil.copy2(archivo_origen, destino)
    print(f"  ✓ Copiado a: exports_flexxus/{destino.name}")

    inyectar_etiqueta(destino, deposito, tipo, fecha_export)

    print(f"\n{'='*60}")
    print(f"  ✅ LISTO")
    print(f"  Archivo: exports_flexxus/{destino.name}")
    print(f"  Depósito: {deposito['full']}")
    print(f"  Tipo:     {tipo['nombre']}")
    print(f"  Fecha:    {fecha_export}")
    print(f"{'='*60}")

    # 8. Preguntar si borrar el original
    borrar = input(f"\n¿Borrar el original de Descargas? (s/N): ").strip().lower()
    if borrar == "s":
        archivo_origen.unlink()
        print(f"  🗑️  Original eliminado.")

    print("\n📂 Carpeta de destino: exports_flexxus/")
    input("\nPresioná Enter para cerrar...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nCancelado.")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        input("\nPresioná Enter para cerrar...")
