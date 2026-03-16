"""
ROKER NEXUS — Importador Order List AI-TECH
Parsea el archivo xlsx de cotización/pedido de Diego (proveedor chino).

Formato conocido:
  - Nombre contiene "AI-TECH" (ej: LA_COTIZACION_DE_AI-TECH_039-20260305.xlsx)
  - Hoja: 'PI'
  - Fila 1: "Invoice" (título)
  - Fila 2: headers (Brand, código, MODELO UNIVERSAL, MODELO STICKER, Spec, -, Quality, Colour, QTY, PRICE, Total)
  - Fila 3: título de sección (ej: "LCD+TOUCH")
  - Filas 4+: datos

Columnas por índice (0-based):
  0: Brand (IPH, M, SA, IN/TE...)
  1: Código proveedor chino (10 dígitos, ej: 2411150023)
  2: MODELO UNIVERSAL (descripción larga)
  3: MODELO STICKER (nombre corto)
  4: Specification (LCD Complete, LCD Complete+Frame...)
  5: Type (AMP, AMM, AMP 2.0...)
  6: Quality (HIGH COPY, OLED, INCELL...)
  7: Colour
  8: QTY
  9: PRICE (USD)
  10: Total (fórmula =QTY*PRICE)
"""
import re
import os
from datetime import datetime
from typing import Optional
import pandas as pd
import openpyxl


def es_orderlist_aitech(filename: str) -> bool:
    """Detecta si un archivo es un Order List de AI-TECH por su nombre."""
    return "AI-TECH" in filename.upper()


def extraer_invoice_numero(filename: str) -> str:
    """
    Extrae el número de invoice del nombre del archivo.
    Ej: LA_COTIZACION_DE_AI-TECH_039-20260305.xlsx → '039'
    """
    nombre = os.path.basename(filename)
    # Buscar patrón NNN después de AI-TECH_
    m = re.search(r'AI-TECH[_-](\d+)', nombre, re.IGNORECASE)
    if m:
        return m.group(1)
    # Alternativa: cualquier secuencia de dígitos en el nombre
    nums = re.findall(r'\d+', nombre)
    if nums:
        # Preferir el que tiene 3 dígitos (número de invoice típico)
        for n in nums:
            if len(n) == 3:
                return n
        return nums[0]
    return "000"


def extraer_fecha_archivo(filename: str) -> str:
    """
    Extrae la fecha del nombre del archivo si la tiene.
    Ej: AI-TECH_039-20260305.xlsx → '2026-03-05'
    """
    nombre = os.path.basename(filename)
    m = re.search(r'(\d{4})(\d{2})(\d{2})', nombre)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return datetime.now().strftime("%Y-%m-%d")


def parsear_orderlist(filepath: str) -> dict:
    """
    Parsea el Order List de AI-TECH y retorna un dict con:
      - invoice_id: str
      - fecha: str
      - filename: str
      - items: list[dict]
      - total_usd: float
      - secciones: list[str]
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Buscar hoja PI (o la primera hoja disponible)
    if 'PI' in wb.sheetnames:
        ws = wb['PI']
    else:
        ws = wb.worksheets[0]

    invoice_id = extraer_invoice_numero(filepath)
    fecha = extraer_fecha_archivo(filepath)
    filename = os.path.basename(filepath)

    items = []
    secciones = []
    seccion_actual = "SIN SECCIÓN"
    total_usd = 0.0

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Fila vacía → saltar
        if not any(v for v in row if v is not None):
            continue

        col0 = str(row[0]).strip() if row[0] else ""
        col1 = row[1]  # código proveedor (numérico o None)
        col2 = str(row[2]).strip() if row[2] else ""  # MODELO UNIVERSAL
        col3 = str(row[3]).strip() if row[3] else ""  # MODELO STICKER
        col4 = str(row[4]).strip() if row[4] else ""  # Specification
        col5 = str(row[5]).strip() if row[5] else ""  # Type
        col6 = str(row[6]).strip() if row[6] else ""  # Quality
        col7 = str(row[7]).strip() if row[7] else ""  # Colour
        col8 = row[8]   # QTY
        col9 = row[9]   # PRICE USD
        # col10 = Total (fórmula, no lo usamos)

        # Detectar fila de sección (col0 con texto, col1 vacío, no es un dato)
        if col0 and not col1 and col0 not in ("Invoice", "Brand:"):
            # Podría ser sección (ej: "LCD+TOUCH", "TOTAL")
            if not col2 or col2 in ("MODELO UNIVERSAL:", ""):
                seccion_actual = col0
                if col0 not in ("TOTAL",) and col0 not in secciones:
                    secciones.append(col0)
                continue

        # Saltar headers y totales
        if col0 in ("Brand:", "Invoice", "TOTAL") or col2 in ("MODELO UNIVERSAL:", ""):
            continue

        # Fila de dato real: debe tener código proveedor (col1) numérico
        try:
            codigo_prov = str(int(col1)) if col1 else None
        except (ValueError, TypeError):
            codigo_prov = str(col1).strip() if col1 else None

        if not codigo_prov and not col2:
            continue

        # Cantidad y precio
        try:
            qty = int(col8) if col8 else 0
        except (ValueError, TypeError):
            qty = 0

        # col9 es PRECIO EN RMB (yuan chino), no USD.
        # Convertir a USD dividiendo por la tasa RMB→USD (por defecto 6.9)
        TASA_RMB_USD = 6.9
        try:
            precio_rmb = float(col9) if col9 else 0.0
        except (ValueError, TypeError):
            precio_rmb = 0.0

        precio = round(precio_rmb / TASA_RMB_USD, 2)
        subtotal = qty * precio
        total_usd += subtotal

        item = {
            "brand":            col0,
            "codigo_proveedor": codigo_prov or "",
            "modelo_universal": col2,
            "modelo_sticker":   col3,
            "specification":    col4,
            "type":             col5,
            "quality":          col6,
            "colour":           col7,
            "cantidad_pedida":  qty,
            "precio_rmb":       precio_rmb,          # precio original en yuan
            "precio_usd":       precio,               # precio_rmb / 6.9
            "subtotal_usd":     subtotal,
            "seccion":          seccion_actual,
            # Campos que se llenarán en el matching
            "codigo_flexxus":   None,
            "descripcion_flexxus": None,
            "match_score":      0,
            "match_confirmado": False,
            "cantidad_recibida": 0,
            "estado_item":      "pendiente",
            "notas":            "",
        }
        items.append(item)

    return {
        "invoice_id":  invoice_id,
        "fecha":       fecha,
        "filename":    filename,
        "proveedor":   "AI-TECH",
        "items":       items,
        "total_usd":   round(total_usd, 2),
        "secciones":   secciones,
        "total_items": len(items),
    }


def hacer_matching_fuzzy(items: list, df_articulos: pd.DataFrame) -> list:
    """
    Para cada item del Order List, busca el código Flexxus más probable.
    Usa el MODELO UNIVERSAL para comparar con descripciones de artículos.

    Retorna los items con campos de matching completados:
      - codigo_flexxus
      - descripcion_flexxus
      - match_score (0-100)
      - match_confirmado (False = necesita confirmación humana)
    """
    from rapidfuzz import fuzz, process

    if df_articulos.empty:
        return items

    # Preparar lista de búsqueda
    opciones = {}
    for _, art in df_articulos.iterrows():
        codigo = str(art.get("codigo", "")).strip()
        desc = str(art.get("descripcion", "")).upper().strip()
        if codigo and desc:
            opciones[codigo] = desc

    codigos = list(opciones.keys())
    descs = list(opciones.values())

    UMBRAL_AUTO = 85    # >= 85 → match automático confirmado
    UMBRAL_SUGERIR = 60  # >= 60 → sugerencia a confirmar

    for item in items:
        query = item["modelo_universal"].upper()
        if not query:
            continue

        # Buscar en descripciones Flexxus
        resultado = process.extractOne(
            query, descs,
            scorer=fuzz.token_set_ratio,
            score_cutoff=UMBRAL_SUGERIR
        )

        if resultado:
            desc_match, score, idx = resultado
            codigo_match = codigos[idx]
            item["codigo_flexxus"] = codigo_match
            item["descripcion_flexxus"] = desc_match
            item["match_score"] = score
            item["match_confirmado"] = score >= UMBRAL_AUTO

    return items


def aplicar_conversion_wf(descripcion: str) -> str:
    """
    Convierte 'C/MARCO' → 'W/F' en descripciones para exportar a Diego.
    Ej: 'MODULO SAM A10S C/MARCO' → 'MODULO SAM A10S W/F'
    """
    conversiones = {
        "C/MARCO": "W/F",
        "CON MARCO": "WITH FRAME",
        "C/ MARCO": "W/F",
    }
    resultado = descripcion.upper()
    for origen, destino in conversiones.items():
        resultado = resultado.replace(origen, destino)
    return resultado


def exportar_para_diego(items: list, invoice_id: str) -> pd.DataFrame:
    """
    Genera el DataFrame para exportar a Diego (proveedor chino).
    Aplica la conversión C/MARCO → W/F en las descripciones.
    """
    rows = []
    for item in items:
        desc_export = aplicar_conversion_wf(item.get("modelo_universal", ""))
        rows.append({
            "Brand":           item.get("brand", ""),
            "Item Code":       item.get("codigo_proveedor", ""),
            "Model":           desc_export,
            "Sticker":         item.get("modelo_sticker", ""),
            "Specification":   item.get("specification", ""),
            "Type":            item.get("type", ""),
            "Quality":         item.get("quality", ""),
            "Colour":          item.get("colour", ""),
            "QTY":             item.get("cantidad_pedida", 0),
            "Unit Price":      item.get("precio_usd", 0),
            "Total":           item.get("subtotal_usd", 0),
        })

    df = pd.DataFrame(rows)
    return df
