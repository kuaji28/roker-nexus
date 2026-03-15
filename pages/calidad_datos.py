"""
ROKER NEXUS — Calidad de Datos
================================
Detecta y gestiona errores de carga en el catálogo Flexxus:
  - Rubro incorrecto (código o descripción sugiere otra marca)
  - Módulos con rubro genérico
  - Duplicados potenciales (descripción muy similar)
  - Códigos huérfanos (en stock pero sin ficha en artículos)
  - Código numérico con referencia a AITECH

Genera reporte de correcciones para enviar a Lorena / Ezequiel / Matías.
"""
import io
from datetime import datetime

import pandas as pd
import streamlit as st

from database import execute_query, query_to_df
from utils.calidad_datos import (
    detectar_errores_calidad,
    guardar_errores_calidad,
    get_errores_pendientes,
)


# ──────────────────────────────────────────────────────────────
#  CONSTANTES UI
# ──────────────────────────────────────────────────────────────

TIPO_LABELS = {
    "RUBRO_INCORRECTO":           "🔴 Rubro incorrecto",
    "RUBRO_GENERICO":             "🟡 Rubro genérico",
    "POSIBLE_DUPLICADO":          "🟠 Posible duplicado",
    "CODIGO_HUERFANO":            "⚫ Código huérfano",
    "CODIGO_PROVEEDOR_INCORRECTO":"🔵 Código proveedor incorrecto",
    "MANUAL":                     "✏️ Ingreso manual",
}

TIPO_OPTIONS = ["Todos"] + list(TIPO_LABELS.keys())

CONFIANZA_ORDER = {"🔴 Alta": 0, "🟡 Media": 1, "⚪ Baja": 2}

STAFF_CORRECTORES = ["Lorena Rodriguez", "Ezequiel Firmapaz", "Matias Toledano"]


# ──────────────────────────────────────────────────────────────
#  DATOS
# ──────────────────────────────────────────────────────────────

def _stock_para_analisis() -> pd.DataFrame:
    """Lee el último snapshot de stock disponible (SAN JOSE primero)."""
    try:
        # Intentar SAN JOSE (SJ) — el más completo
        df = query_to_df("""
            SELECT codigo, descripcion, rubro, deposito, fecha
            FROM stock_snapshots
            WHERE deposito = 'SJ'
            ORDER BY fecha DESC
            LIMIT 1
        """)
        if df.empty:
            # Fallback: cualquier snapshot disponible
            df = query_to_df("""
                SELECT s.codigo, s.descripcion, s.rubro, s.deposito, s.fecha
                FROM stock_snapshots s
                INNER JOIN (
                    SELECT deposito, MAX(fecha) AS max_fecha
                    FROM stock_snapshots
                    GROUP BY deposito
                ) latest ON s.deposito = latest.deposito AND s.fecha = latest.max_fecha
            """)
        return df
    except Exception:
        return pd.DataFrame()


def _stock_completo_ultimas_fechas() -> pd.DataFrame:
    """Snapshot completo más reciente de cada depósito, unificado."""
    try:
        return query_to_df("""
            SELECT s.codigo, s.descripcion, s.rubro, s.deposito
            FROM stock_snapshots s
            INNER JOIN (
                SELECT deposito, MAX(fecha) AS max_fecha
                FROM stock_snapshots
                GROUP BY deposito
            ) latest ON s.deposito = latest.deposito AND s.fecha = latest.max_fecha
        """)
    except Exception:
        return pd.DataFrame()


def _articulos_df() -> pd.DataFrame:
    try:
        return query_to_df("SELECT codigo FROM articulos")
    except Exception:
        return pd.DataFrame()


def _summary_stats() -> dict:
    """Cuenta rápida de anomalías abiertas por tipo."""
    try:
        rows = execute_query("""
            SELECT tipo, COUNT(*) as cnt
            FROM anomalias
            WHERE estado = 'abierta'
            GROUP BY tipo
        """, fetch=True)
        stats = {r["tipo"]: r["cnt"] for r in (rows or [])}
        stats["_total"] = sum(stats.values())
        return stats
    except Exception:
        return {"_total": 0}


def _cerrar_anomalia(anomalia_id: int, notas_cierre: str = ""):
    try:
        execute_query(
            "UPDATE anomalias SET estado='resuelta', notas=notas||? WHERE id=?",
            (f" | CIERRE: {notas_cierre}" if notas_cierre else "", anomalia_id),
            fetch=False,
        )
    except Exception:
        pass


def _insertar_anomalia_manual(codigo: str, tipo: str, descripcion: str,
                               rubro_actual: str, rubro_sugerido: str,
                               severidad: str, notas: str):
    try:
        existe = execute_query(
            "SELECT 1 FROM anomalias WHERE codigo=? AND tipo=? AND estado='abierta'",
            (codigo.upper(), tipo), fetch=True,
        )
        if existe:
            return False
        execute_query("""
            INSERT INTO anomalias (codigo, tipo, descripcion, severidad, estado, detectada_en, notas)
            VALUES (?, ?, ?, ?, 'abierta', ?, ?)
        """, (
            codigo.upper(), tipo,
            f"{descripcion} | Rubro actual: {rubro_actual} | Sugerido: {rubro_sugerido}",
            severidad,
            datetime.now().isoformat(),
            f"[manual] {notas}",
        ), fetch=False)
        return True
    except Exception:
        return False


# ──────────────────────────────────────────────────────────────
#  ANÁLISIS: CORRER DETECCIÓN
# ──────────────────────────────────────────────────────────────

def _correr_analisis() -> dict:
    """Ejecuta el motor de calidad sobre el stock actual y persiste en DB."""
    df_stock = _stock_completo_ultimas_fechas()
    if df_stock.empty:
        return {"ok": False, "msg": "No hay datos de stock cargados. Importá al menos un archivo de stock primero."}

    errores = detectar_errores_calidad(df_stock)

    # Detectar huérfanos si hay catálogo de artículos
    try:
        from utils.calidad_datos import detectar_huerfanos
        df_arts = _articulos_df()
        if not df_arts.empty:
            huerfanos = detectar_huerfanos(df_stock, df_arts)
            errores.extend(huerfanos)
    except Exception:
        pass

    nuevos = guardar_errores_calidad(errores, fuente="analisis_manual")

    return {
        "ok": True,
        "total_detectados": len(errores),
        "nuevos_en_db": nuevos,
        "depositos_analizados": df_stock["deposito"].nunique() if "deposito" in df_stock.columns else 1,
        "items_analizados": len(df_stock),
    }


# ──────────────────────────────────────────────────────────────
#  EXPORTAR REPORTE EXCEL
# ──────────────────────────────────────────────────────────────

def _generar_excel_reporte(df_errores: pd.DataFrame, destinatario: str) -> bytes:
    """Genera el reporte de correcciones formateado para el staff."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Correcciones Flexxus"

        # ── Paleta de colores ──
        COLOR_HEADER_BG  = "1F3864"   # azul oscuro
        COLOR_HEADER_FG  = "FFFFFF"
        COLOR_ALTA_BG    = "FFE0E0"   # rojo suave
        COLOR_MEDIA_BG   = "FFF3CD"   # amarillo suave
        COLOR_BAJA_BG    = "F0F0F0"   # gris
        COLOR_ALT_ROW    = "F7F9FC"

        thin_side = Side(style="thin", color="CCCCCC")
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)

        # ── Fila 1: título principal ──
        ws.merge_cells("A1:G1")
        ws["A1"] = f"ROKER NEXUS — Reporte de Calidad de Datos Flexxus"
        ws["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER_FG)
        ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Fila 2: meta info ──
        ws.merge_cells("A2:G2")
        fecha_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws["A2"] = f"Generado: {fecha_str}  |  Para: {destinatario}  |  Total errores: {len(df_errores)}"
        ws["A2"].font = Font(italic=True, size=10, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center")

        # ── Fila 3: separador vacío ──
        ws.row_dimensions[3].height = 6

        # ── Fila 4: headers de tabla ──
        HEADERS = ["Código", "Descripción", "Rubro Actual", "Rubro Sugerido",
                   "Tipo de Error", "Confianza", "Qué hacer"]
        header_row = 4
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=10)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 20

        # ── Datos ──
        col_map = {
            "codigo": 1, "descripcion": 2, "rubro_actual": 3,
            "rubro_sugerido": 4, "tipo_error": 5,
            "confianza": 6, "sugerencia_correccion": 7,
        }

        # Reordenar por confianza (alta primero)
        df_sorted = df_errores.copy()
        if "confianza" in df_sorted.columns:
            df_sorted["_orden"] = df_sorted["confianza"].map(
                lambda x: 0 if "Alta" in str(x) else (1 if "Media" in str(x) else 2)
            )
            df_sorted = df_sorted.sort_values("_orden").drop(columns=["_orden"])

        for row_num, (_, row) in enumerate(df_sorted.iterrows(), start=header_row + 1):
            # Color de fila por confianza
            conf = str(row.get("confianza", ""))
            if "Alta" in conf:
                row_bg = COLOR_ALTA_BG
            elif "Media" in conf:
                row_bg = COLOR_MEDIA_BG
            else:
                row_bg = COLOR_BAJA_BG if row_num % 2 == 0 else COLOR_ALT_ROW

            fill = PatternFill("solid", fgColor=row_bg)

            for col_key, col_idx in col_map.items():
                val = row.get(col_key, "—")
                # Limpiar emojis de confianza para que quede limpio
                if col_key == "confianza":
                    val = str(val).replace("🔴 ", "🔴 ").replace("🟡 ", "").replace("⚪ ", "")
                cell = ws.cell(row=row_num, column=col_idx, value=str(val) if val else "—")
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=9)
            ws.row_dimensions[row_num].height = 32

        # ── Anchos de columna ──
        COL_WIDTHS = [18, 42, 22, 22, 28, 14, 48]
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Freeze panes (header siempre visible) ──
        ws.freeze_panes = ws["A5"]

        # ── Guardar en memoria ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except Exception as e:
        st.error(f"Error generando Excel: {e}")
        return b""


# ──────────────────────────────────────────────────────────────
#  RENDER PRINCIPAL
# ──────────────────────────────────────────────────────────────

def render():
    st.title("🔍 Calidad de Datos")
    st.caption("Detecta errores de carga en Flexxus — rubros incorrectos, duplicados, huérfanos.")

    # ── Session state ──
    st.session_state.setdefault("calidad_tab", "errores")
    st.session_state.setdefault("calidad_filtro_tipo", "Todos")
    st.session_state.setdefault("calidad_ultimo_analisis", None)

    # ── KPIs resumen ──
    stats = _summary_stats()
    total = stats.get("_total", 0)

    col_a, col_b, col_c, col_d, col_e = st.columns(5)
    with col_a:
        st.metric("Total abiertos", total, border=True)
    with col_b:
        n = stats.get("RUBRO_INCORRECTO", 0)
        st.metric("🔴 Rubro incorrecto", n, border=True)
    with col_c:
        n = stats.get("RUBRO_GENERICO", 0) + stats.get("CODIGO_PROVEEDOR_INCORRECTO", 0)
        st.metric("🟡 Otros errores", n, border=True)
    with col_d:
        n = stats.get("POSIBLE_DUPLICADO", 0)
        st.metric("🟠 Duplicados", n, border=True)
    with col_e:
        n = stats.get("CODIGO_HUERFANO", 0)
        st.metric("⚫ Huérfanos", n, border=True)

    st.markdown("---")

    # ── Barra de acciones ──
    col_btn1, col_btn2, col_spacer = st.columns([2, 2, 6])

    with col_btn1:
        if st.button("🔄 Analizar ahora", type="primary", use_container_width=True,
                     help="Corre el motor de detección sobre el stock actual en la DB"):
            with st.spinner("Analizando datos..."):
                resultado = _correr_analisis()
            if resultado["ok"]:
                st.session_state["calidad_ultimo_analisis"] = datetime.now().strftime("%H:%M:%S")
                st.success(
                    f"✅ Análisis completado — "
                    f"{resultado['items_analizados']:,} ítems · "
                    f"{resultado['total_detectados']} errores detectados · "
                    f"**{resultado['nuevos_en_db']} nuevos** en la DB"
                )
                st.rerun()
            else:
                st.warning(f"⚠️ {resultado['msg']}")

    with col_btn2:
        ultimo = st.session_state.get("calidad_ultimo_analisis")
        if ultimo:
            st.caption(f"Último análisis: {ultimo}")

    # ── Tabs ──
    tab_errores, tab_manual, tab_resueltos = st.tabs([
        "📋 Errores detectados",
        "✏️ Agregar manual",
        "✅ Resueltos",
    ])

    # ══════════════════════════════════════════════════════
    # TAB 1 — Errores detectados
    # ══════════════════════════════════════════════════════
    with tab_errores:
        df_errores = get_errores_pendientes()

        if df_errores.empty:
            st.info("No hay errores abiertos. Hacé clic en **Analizar ahora** para correr el motor de detección.")
        else:
            # Filtros
            col_f1, col_f2, col_f3 = st.columns([3, 3, 6])
            with col_f1:
                tipo_sel = st.selectbox(
                    "Filtrar por tipo",
                    options=TIPO_OPTIONS,
                    format_func=lambda x: TIPO_LABELS.get(x, x) if x != "Todos" else "Todos los tipos",
                    key="calidad_filtro_tipo_sel",
                )
            with col_f2:
                buscar = st.text_input("Buscar código o descripción", placeholder="Ej: MSAMS...")

            # Aplicar filtros
            df_view = df_errores.copy()
            if tipo_sel != "Todos":
                df_view = df_view[df_view["tipo_error"] == tipo_sel]
            if buscar:
                mask = (
                    df_view["codigo"].astype(str).str.upper().str.contains(buscar.upper(), na=False) |
                    df_view["descripcion"].astype(str).str.upper().str.contains(buscar.upper(), na=False)
                )
                df_view = df_view[mask]

            # Stats post-filtro
            st.caption(f"Mostrando **{len(df_view)}** de {len(df_errores)} errores abiertos")

            # Tabla principal
            cols_show = [c for c in ["codigo", "tipo_error", "descripcion", "confianza", "regla", "detectada_en"]
                         if c in df_view.columns]
            st.dataframe(
                df_view[cols_show] if cols_show else df_view,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "codigo":        st.column_config.TextColumn("Código", width=130, pinned=True),
                    "tipo_error":    st.column_config.TextColumn("Tipo", width=200),
                    "descripcion":   st.column_config.TextColumn("Descripción / Detalle", width=340),
                    "confianza":     st.column_config.TextColumn("Confianza", width=120),
                    "regla":         st.column_config.TextColumn("Regla disparada", width=300),
                    "detectada_en":  st.column_config.DatetimeColumn("Detectado", format="DD/MM/YYYY HH:mm", width=140),
                },
                height=380,
            )

            # ── Exportar reporte ──
            st.markdown("#### 📤 Exportar reporte de correcciones")
            col_dest, col_exp = st.columns([3, 2])
            with col_dest:
                destinatario = st.selectbox(
                    "Para",
                    options=STAFF_CORRECTORES + ["Todos (cc: Lorena, Ezequiel, Matías)"],
                    index=0,
                )
            with col_exp:
                st.write("")  # spacer
                st.write("")
                excel_bytes = _generar_excel_reporte(df_view, destinatario)
                fecha_archivo = datetime.now().strftime("%Y%m%d")
                nombre_archivo = f"calidad_flexxus_{fecha_archivo}.xlsx"
                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_bytes,
                    file_name=nombre_archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                    disabled=not excel_bytes,
                )

            # ── Resolver errores ──
            with st.expander("✅ Marcar errores como resueltos", expanded=False):
                st.caption("Ingresá el ID de la anomalía y una nota de cierre.")
                ids_disponibles = []
                if "id" in df_view.columns:
                    ids_disponibles = df_view["id"].dropna().astype(int).tolist()

                if ids_disponibles:
                    col_id, col_nota, col_res = st.columns([2, 5, 2])
                    with col_id:
                        id_resolver = st.number_input(
                            "ID anomalía",
                            min_value=1,
                            step=1,
                            value=ids_disponibles[0] if ids_disponibles else 1,
                        )
                    with col_nota:
                        nota_cierre = st.text_input("Nota de cierre (opcional)",
                                                     placeholder="Ej: Corregido por Lorena el 15/03")
                    with col_res:
                        st.write("")
                        st.write("")
                        if st.button("Resolver", use_container_width=True):
                            _cerrar_anomalia(id_resolver, nota_cierre)
                            st.success(f"Anomalía #{id_resolver} marcada como resuelta.")
                            st.rerun()
                else:
                    st.info("La tabla no incluye columna de ID. Accedé directo desde la DB para resolver.")

    # ══════════════════════════════════════════════════════
    # TAB 2 — Agregar manual
    # ══════════════════════════════════════════════════════
    with tab_manual:
        st.markdown("#### ✏️ Registrar error detectado manualmente")
        st.caption("Para errores que el motor automático no detecta (Ej: descripción con typo, precio atípico, etc.)")

        with st.form("form_manual_error", clear_on_submit=True):
            col_m1, col_m2 = st.columns(2)
            with col_m1:
                m_codigo = st.text_input("Código del artículo *", placeholder="Ej: MSAMA02S")
                m_rubro_actual = st.text_input("Rubro actual en Flexxus *",
                                                placeholder="Ej: MODULOS")
            with col_m2:
                m_tipo = st.selectbox(
                    "Tipo de error *",
                    options=[k for k in TIPO_LABELS if k != "MANUAL"],
                    format_func=lambda x: TIPO_LABELS.get(x, x),
                )
                m_rubro_sugerido = st.text_input("Rubro sugerido",
                                                  placeholder="Ej: SAMSUNG")

            m_descripcion = st.text_area("Descripción del problema *",
                                          placeholder="Ej: Módulo Samsung A32 cargado bajo rubro MODULOS en vez de SAMSUNG",
                                          height=80)
            m_severidad = st.radio("Severidad", options=["alta", "media", "baja"],
                                    horizontal=True, index=1)
            m_notas = st.text_input("Notas adicionales (opcional)")

            submitted = st.form_submit_button("Agregar error", type="primary")

        if submitted:
            if not m_codigo or not m_descripcion or not m_rubro_actual:
                st.error("Completá los campos obligatorios (Código, Rubro actual, Descripción).")
            else:
                ok = _insertar_anomalia_manual(
                    codigo=m_codigo,
                    tipo=m_tipo,
                    descripcion=m_descripcion,
                    rubro_actual=m_rubro_actual,
                    rubro_sugerido=m_rubro_sugerido,
                    severidad=m_severidad,
                    notas=m_notas,
                )
                if ok:
                    st.success(f"✅ Error para {m_codigo.upper()} registrado correctamente.")
                    st.rerun()
                else:
                    st.warning(f"Ya existe una anomalía abierta de tipo '{m_tipo}' para el código '{m_codigo.upper()}'. "
                               "Revisá la pestaña Errores detectados.")

    # ══════════════════════════════════════════════════════
    # TAB 3 — Resueltos
    # ══════════════════════════════════════════════════════
    with tab_resueltos:
        st.markdown("#### ✅ Historial de errores resueltos")
        try:
            df_res = query_to_df("""
                SELECT codigo, tipo as tipo_error, descripcion, severidad,
                       detectada_en, notas
                FROM anomalias
                WHERE tipo IN (
                    'RUBRO_INCORRECTO','RUBRO_GENERICO','POSIBLE_DUPLICADO',
                    'CODIGO_HUERFANO','CODIGO_PROVEEDOR_INCORRECTO','MANUAL'
                )
                AND estado = 'resuelta'
                ORDER BY detectada_en DESC
                LIMIT 200
            """)
        except Exception:
            df_res = pd.DataFrame()

        if df_res.empty:
            st.info("Aún no hay errores marcados como resueltos.")
        else:
            st.caption(f"{len(df_res)} errores resueltos")
            st.dataframe(
                df_res,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "codigo":       st.column_config.TextColumn("Código", width=130),
                    "tipo_error":   st.column_config.TextColumn("Tipo", width=200),
                    "descripcion":  st.column_config.TextColumn("Descripción", width=340),
                    "severidad":    st.column_config.TextColumn("Severidad", width=90),
                    "detectada_en": st.column_config.DatetimeColumn("Detectado", format="DD/MM/YYYY", width=120),
                    "notas":        st.column_config.TextColumn("Notas cierre", width=300),
                },
                height=340,
            )

            # Estadística rápida de efectividad
            col_s1, col_s2 = st.columns(2)
            with col_s1:
                por_tipo = df_res["tipo_error"].value_counts()
                st.markdown("**Por tipo:**")
                for tipo, cnt in por_tipo.items():
                    label = TIPO_LABELS.get(tipo, tipo)
                    st.caption(f"{label}: {cnt}")


# ──────────────────────────────────────────────────────────────
#  ENTRY POINT
# ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    render()
